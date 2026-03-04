import os
import time
import shutil
import pandas as pd
import argparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options

# --- Scraper Setup ---
DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

def select_primefaces_dropdown(browser, wait, container_id, option_text):
    container = wait.until(EC.element_to_be_clickable((By.ID, container_id)))
    container.click()
    panel_id = container_id + '_panel'
    panel = wait.until(EC.visibility_of_element_located((By.ID, panel_id)))
    options = panel.find_elements(By.TAG_NAME, 'li')
    for opt in options:
        if opt.text.strip() == option_text:
            opt.click()
            time.sleep(0.5)
            return
    raise Exception(f'Option "{option_text}" not found in {container_id}')

def wait_for_yaxis_change(browser, expected):
    for _ in range(20):
        label = browser.find_element(By.ID, "yaxisVar_label").text.strip()
        if label == expected:
            return True
        time.sleep(0.3)
    raise Exception(f"Y-Axis did not change to {expected}")

def wait_for_download(filename, timeout=30):
    for _ in range(timeout * 2):
        files = [f for f in os.listdir(DOWNLOAD_DIR) if f.startswith("reportTable") and f.endswith(".xlsx")]
        if files:
            return sorted(files, key=lambda x: os.path.getmtime(os.path.join(DOWNLOAD_DIR, x)))[-1]
        time.sleep(0.5)
    raise Exception("Download did not complete in time")

def wait_for_overlay_to_disappear(browser):
    for _ in range(20):
        overlay = browser.find_element(By.ID, "j_idt135_blocker")
        if overlay.value_of_css_property("display") == "none":
            return
        time.sleep(0.3)

# --- Scraping ---
def run_scraper(selected_types=None, start_year=2018, end_year=2026):
    edge_options = Options()
    edge_options.use_chromium = True
    edge_options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })
    browser = webdriver.Edge(options=edge_options)
    url = "https://vahan.parivahan.gov.in/vahan4dashboard/vahan/view/reportview.xhtml"
    browser.get(url)
    wait = WebDriverWait(browser, 30)
    time.sleep(5)
    select_primefaces_dropdown(browser, wait, "xaxisVar", "Month Wise")
    year_type_container = wait.until(EC.presence_of_element_located((By.ID, "selectedYearType")))
    if "ui-state-disabled" not in year_type_container.get_attribute("class"):
        select_primefaces_dropdown(browser, wait, "selectedYearType", "Calendar Year")
    yaxis_options = [
        "Vehicle Category",
        "Vehicle Class",
        "Norms",
        "Fuel",
        "Maker",
        "State"
    ]
    if selected_types:
        yaxis_options = [y for y in yaxis_options if y in selected_types]
    for yaxis in yaxis_options:
        for attempt_yaxis in range(3):
            try:
                select_primefaces_dropdown(browser, wait, "yaxisVar", yaxis)
                wait_for_yaxis_change(browser, yaxis)
                break
            except Exception as e:
                print(f"Retrying Y-Axis {yaxis} (attempt {attempt_yaxis+1}): {e}")
                time.sleep(2)
        yaxis_folder = os.path.join(DOWNLOAD_DIR, yaxis.replace(" ", "_"))
        os.makedirs(yaxis_folder, exist_ok=True)
        for year in range(start_year, end_year + 1):
            success = False
            for attempt in range(3):
                try:
                    # Always re-select Y-Axis and Year to get fresh elements
                    select_primefaces_dropdown(browser, wait, "yaxisVar", yaxis)
                    wait_for_yaxis_change(browser, yaxis)
                    select_primefaces_dropdown(browser, wait, "selectedYear", str(year))
                    refresh_btn = wait.until(EC.element_to_be_clickable((By.ID, "j_idt73")))
                    refresh_btn.click()
                    time.sleep(2)
                    export_btn = wait.until(EC.element_to_be_clickable((By.ID, "groupingTable:xls")))
                    wait_for_overlay_to_disappear(browser)
                    export_btn.click()
                    downloaded = wait_for_download("reportTable.xlsx")
                    new_name = f"{yaxis.replace(' ', '_')}_{year}.xlsx"
                    shutil.move(os.path.join(DOWNLOAD_DIR, downloaded), os.path.join(yaxis_folder, new_name))
                    print(f"Saved: {os.path.join(yaxis_folder, new_name)}")
                    success = True
                    break
                except Exception as e:
                    print(f"Retry {attempt+1}/3 for {yaxis} {year}: {e}")
                    if attempt < 2:
                        time.sleep(2)
            if not success:
                print(f"All attempts failed for {yaxis} {year}. Reloading page and retrying once.")
                browser.refresh()
                time.sleep(5)
                try:
                    select_primefaces_dropdown(browser, wait, "xaxisVar", "Month Wise")
                    year_type_container = wait.until(EC.presence_of_element_located((By.ID, "selectedYearType")))
                    if "ui-state-disabled" not in year_type_container.get_attribute("class"):
                        select_primefaces_dropdown(browser, wait, "selectedYearType", "Calendar Year")
                    select_primefaces_dropdown(browser, wait, "yaxisVar", yaxis)
                    wait_for_yaxis_change(browser, yaxis)
                    select_primefaces_dropdown(browser, wait, "selectedYear", str(year))
                    refresh_btn = wait.until(EC.element_to_be_clickable((By.ID, "j_idt73")))
                    refresh_btn.click()
                    time.sleep(2)
                    export_btn = wait.until(EC.element_to_be_clickable((By.ID, "groupingTable:xls")))
                    wait_for_overlay_to_disappear(browser)
                    export_btn.click()
                    downloaded = wait_for_download("reportTable.xlsx")
                    new_name = f"{yaxis.replace(' ', '_')}_{year}.xlsx"
                    shutil.move(os.path.join(DOWNLOAD_DIR, downloaded), os.path.join(yaxis_folder, new_name))
                    print(f"Saved after reload: {os.path.join(yaxis_folder, new_name)}")
                except Exception as e:
                    print(f"Final retry after reload failed for {yaxis} {year}: {e}")
    browser.quit()

# --- Master Sheet Compilation ---
def compile_master_sheet(clean_master=False):
    criteria_folders = [f for f in os.listdir(DOWNLOAD_DIR) if os.path.isdir(os.path.join(DOWNLOAD_DIR, f))]
    writer = pd.ExcelWriter(os.path.join(DOWNLOAD_DIR, 'master_sheet.xlsx'), engine='openpyxl')
    for folder in criteria_folders:
        folder_path = os.path.join(DOWNLOAD_DIR, folder)
        year_files = sorted([f for f in os.listdir(folder_path) if f.endswith('.xlsx')], key=lambda x: int(x.split('_')[-1].replace('.xlsx','')))
        dfs = []
        keys_set = set()
        headers = []
        for file in year_files:
            file_path = os.path.join(folder_path, file)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                cell_a1 = ws['A1'].value
                # Copy cell A1 to B4
                ws['B4'] = cell_a1
                wb.save(file_path)
                wb.close()
                df = pd.read_excel(file_path, header=None)
                # Remove first 3 rows and column 1
                df = df.iloc[3:].reset_index(drop=True)
                df = df.drop(df.columns[0], axis=1)
                # First row is header
                header_row = df.iloc[0].fillna('')
                # Ensure all columns have unique names (keep empty/nan columns)
                header_row = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(header_row)]
                df.columns = header_row
                df = df[1:].reset_index(drop=True)
                # Use column 1 (now first col) as key
                df = df.set_index(df.columns[0])
                keys_set.update(df.index)
                dfs.append(df)
                headers.append(header_row[1:])
            except Exception as e:
                print(f"Failed to read {file_path}: {e}")
        # Merge all dfs horizontally by index (key)
        if dfs:
            all_keys = sorted(keys_set)
            merged = pd.DataFrame(index=all_keys)
            for i, df in enumerate(dfs):
                year = int(year_files[i].split('_')[-1].replace('.xlsx', ''))
                df.index.name = None
                # Column names: "Jan 2018", "Feb 2018", ... or "Total 2018" for non-month 13th column
                month_names_raw = ('JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC')
                def month_year_label(col):
                    col_str = str(col).strip().upper()
                    if len(col_str) >= 3 and col_str[:3] in month_names_raw:
                        month_str = str(col).strip()
                        month_display = month_str[:1].upper() + month_str[1:].lower() if len(month_str) >= 3 else month_str
                        return f"{month_display} {year}"
                    return f"Total {year}"  # 13th column (e.g. Col_12, Total) → Total YYYY
                df_cols = [month_year_label(col) for col in df.columns if col != df.index.name]
                df_to_join = df.loc[:, [col for col in df.columns if col != df.index.name]]
                df_to_join.columns = df_cols
                merged = merged.join(df_to_join, how='outer')
            merged = merged.reset_index()
            merged_df = merged.rename(columns={'index': 'Key'})
            # Only 1st column (Key) is non-numeric; parse all other columns as int (strip commas first)
            key_col_name = merged_df.columns[0]
            numeric_cols = [c for c in merged_df.columns if c != key_col_name]
            def parse_numeric_series(s):
                return pd.to_numeric(s.astype(str).str.replace(',', '', regex=False), errors='coerce')
            merged_df[numeric_cols] = merged_df[numeric_cols].apply(parse_numeric_series).fillna(0).astype('int64')
            # Add Total (Year) only if not already present (source may have 13th column as Total)
            month_names = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
            years_in_data = set()
            for c in numeric_cols:
                parts = str(c).strip().split()
                if len(parts) == 2 and parts[0] in month_names and parts[1].isdigit():
                    years_in_data.add(int(parts[1]))
            for year in sorted(years_in_data):
                if f'Total {year}' in merged_df.columns:
                    continue
                year_cols = [c for c in numeric_cols if str(c).strip().endswith(str(year)) and str(c).split()[0] in month_names]
                if len(year_cols) == 12:
                    merged_df[f'Total {year}'] = merged_df[year_cols].sum(axis=1)
            # Put Total YYYY columns at end (after all month columns)
            all_cols = [c for c in merged_df.columns if c != key_col_name]
            month_cols_ordered = [c for c in all_cols if not str(c).startswith('Total ')]
            total_cols = [c for c in all_cols if str(c).startswith('Total ')]
            merged_df = merged_df[[key_col_name] + month_cols_ordered + total_cols]
            sheet_name = folder.replace('_', ' ')[:31]
            merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.close()
    print('Master sheet created at:', os.path.join(DOWNLOAD_DIR, 'master_sheet.xlsx'))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Vahan Dashboard Scraper")
    parser.add_argument('--type', nargs='*', help='Y-Axis types to scrape (e.g. "Maker" "Fuel")')
    parser.add_argument('--start-year', type=int, default=2018, help='Start year (inclusive)')
    parser.add_argument('--end-year', type=int, default=2026, help='End year (inclusive)')
    parser.add_argument('--scrape', action='store_true', help='Run scraping only (no compilation)')
    parser.add_argument('--compile', action='store_true', help='Compile master sheet only (no scraping)')
    parser.add_argument('--clean-master', action='store_true', help='Clean master sheet: remove first 3 rows and rename columns to month-year format')
    parser.add_argument('--all', action='store_true', help='Run full pipeline: scrape, compile, and clean master sheet')
    args = parser.parse_args()

    # Default: run full pipeline if no specific action is given
    if args.all or (not args.scrape and not args.compile and not args.clean_master):
        run_scraper(selected_types=args.type, start_year=args.start_year, end_year=args.end_year)
        compile_master_sheet(clean_master=True)
    elif args.scrape:
        run_scraper(selected_types=args.type, start_year=args.start_year, end_year=args.end_year)
    elif args.compile:
        compile_master_sheet(clean_master=args.clean_master)
    elif args.clean_master:
        # Only clean the master sheet (assumes master_sheet.xlsx already exists)
        compile_master_sheet(clean_master=True)
